#!/usr/bin/env python3
"""Profile CSV/XLSX datasets for interactive analysis sessions."""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import math
from collections import Counter
from datetime import datetime
from pathlib import Path
from typing import Any

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover - optional dependency
    load_workbook = None


DATE_HINTS = ("date", "time", "day", "week", "month", "year", "dt")
ID_HINTS = (
    "id",
    "uuid",
    "code",
    "no",
    "编号",
    "订单号",
    "单号",
    "卡号",
    "票号",
    "行号",
)

BUSINESS_FIELD_HINTS = (
    ("利润", ("profit", "利润", "毛利", "净利")),
    ("收入", ("revenue", "income", "收入", "营收")),
    ("销售额", ("sales", "销售额", "销售金额", "成交额")),
    ("费用", ("expense", "cost", "费用", "成本", "支出")),
    ("预算", ("budget", "预算")),
    ("订单量", ("order_count", "orders", "订单量", "订单数")),
    ("客户", ("customer", "client", "客户")),
    ("产品", ("product", "sku", "商品", "产品")),
    ("部门", ("department", "dept", "部门")),
    ("员工", ("employee", "staff", "员工")),
    ("供应商", ("supplier", "vendor", "供应商")),
    ("区域", ("region", "area", "地区", "区域")),
    ("日期", ("date", "日期")),
    ("月份", ("month", "月份", "月度")),
    ("季度", ("quarter", "季度")),
    ("年度", ("year", "年度", "年份")),
)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("dataset", nargs="+", help="Path to one or more CSV/XLSX files")
    parser.add_argument("--sheet", help="Worksheet name for XLSX input")
    parser.add_argument(
        "--all-sheets",
        action="store_true",
        help="Profile every visible worksheet and emit a discovery payload.",
    )
    parser.add_argument("--output", help="Write JSON output to a file")
    return parser.parse_args()


def parse_number(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        if isinstance(value, float) and math.isnan(value):
            return None
        return float(value)
    text = str(value).strip()
    if not text:
        return None
    if text.endswith("%"):
        text = text[:-1]
    text = text.replace(",", "")
    try:
        return float(text)
    except ValueError:
        return None


def parse_date(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.isoformat()
    text = str(value).strip()
    if not text:
        return None
    if text.isdigit():
        if len(text) not in {6, 8}:
            return None
        if not text.startswith(("19", "20")):
            return None
    elif not any(separator in text for separator in ("-", "/", ":", "T", " ")):
        return None
    formats = (
        "%Y-%m-%d",
        "%Y/%m/%d",
        "%Y-%m",
        "%Y/%m",
        "%Y%m%d",
        "%Y%m",
        "%m/%d/%Y",
        "%d/%m/%Y",
        "%Y-%m-%d %H:%M:%S",
    )
    for fmt in formats:
        try:
            return datetime.strptime(text, fmt).isoformat()
        except ValueError:
            continue
    return None


def load_rows(path: Path, sheet_name: str | None = None) -> list[dict[str, Any]]:
    suffix = path.suffix.lower()
    if suffix in {".csv", ".tsv"}:
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            sample = handle.read(4096)
            handle.seek(0)
            dialect = csv.Sniffer().sniff(sample, delimiters=",\t;")
            reader = csv.DictReader(handle, dialect=dialect)
            return [dict(row) for row in reader]
    if suffix == ".xlsx":
        if load_workbook is None:
            raise SystemExit("openpyxl is required to profile .xlsx files")
        workbook = load_workbook(path, read_only=True, data_only=True)
        worksheet = workbook[sheet_name] if sheet_name else workbook.active
        rows = list(worksheet.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(cell).strip() if cell is not None else f"column_{idx+1}" for idx, cell in enumerate(rows[0])]
        result = []
        for row in rows[1:]:
            result.append({headers[idx]: row[idx] if idx < len(row) else None for idx in range(len(headers))})
        return result
    raise SystemExit(f"Unsupported dataset format: {path.suffix}")


def summarize_column(name: str, values: list[Any]) -> dict[str, Any]:
    non_empty = [value for value in values if value not in (None, "")]
    missing_count = len(values) - len(non_empty)
    numeric_values = [number for number in (parse_number(value) for value in non_empty) if number is not None]
    date_values = [date for date in (parse_date(value) for value in non_empty) if date is not None]
    sample_values = []
    seen = set()
    for value in non_empty:
        text = str(value)
        if text not in seen:
            sample_values.append(text)
            seen.add(text)
        if len(sample_values) >= 5:
            break

    distinct_count = len(set(map(str, non_empty))) if non_empty else 0
    numeric_ratio = len(numeric_values) / len(non_empty) if non_empty else 0
    date_ratio = len(date_values) / len(non_empty) if non_empty else 0

    inferred_role = "dimension"
    lowered = name.lower()
    if numeric_ratio >= 0.8 and not any(hint in lowered for hint in ID_HINTS):
        inferred_role = "metric"
    elif date_ratio >= 0.8 or any(hint in lowered for hint in DATE_HINTS):
        inferred_role = "time"
    elif any(hint in lowered for hint in ID_HINTS):
        inferred_role = "identifier"

    summary = {
        "name": name,
        "missing_count": missing_count,
        "missing_rate": round(missing_count / len(values), 4) if values else 0,
        "distinct_count": distinct_count,
        "sample_values": sample_values,
        "numeric_ratio": round(numeric_ratio, 4),
        "date_ratio": round(date_ratio, 4),
        "inferred_role": inferred_role,
    }
    if numeric_values:
        ordered = sorted(numeric_values)
        summary["numeric_summary"] = {
            "min": ordered[0],
            "max": ordered[-1],
            "mean": round(sum(ordered) / len(ordered), 4),
            "zero_rate": round(sum(1 for value in ordered if value == 0) / len(ordered), 4),
        }
    if date_values:
        ordered = sorted(date_values)
        summary["date_summary"] = {"min": ordered[0], "max": ordered[-1]}
    return summary


def build_profile(path: Path, rows: list[dict[str, Any]]) -> dict[str, Any]:
    columns = list(rows[0].keys()) if rows else []
    column_summaries = []
    for column in columns:
        column_summaries.append(summarize_column(column, [row.get(column) for row in rows]))

    roles = Counter(summary["inferred_role"] for summary in column_summaries)
    return {
        "dataset_path": str(path),
        "row_count": len(rows),
        "column_count": len(columns),
        "columns": column_summaries,
        "candidate_time_fields": [summary["name"] for summary in column_summaries if summary["inferred_role"] == "time"],
        "candidate_metric_fields": [summary["name"] for summary in column_summaries if summary["inferred_role"] == "metric"],
        "candidate_dimension_fields": [summary["name"] for summary in column_summaries if summary["inferred_role"] == "dimension"],
        "role_counts": dict(roles),
    }


def workbook_sheets(path: Path) -> list[str | None]:
    if path.suffix.lower() != ".xlsx":
        return [None]
    if load_workbook is None:
        raise SystemExit("openpyxl is required to profile .xlsx files")
    workbook = load_workbook(path, read_only=True, data_only=True)
    return [worksheet.title for worksheet in workbook.worksheets if worksheet.sheet_state == "visible"]


def normalize_field_name(name: str) -> str:
    return "".join(character.lower() for character in name if character.isalnum())


def contract_id(payload: dict[str, Any]) -> str:
    serialized = json.dumps(payload, ensure_ascii=False, sort_keys=True).encode("utf-8")
    return hashlib.sha256(serialized).hexdigest()[:12]


def discovery_fields(items: list[dict[str, Any]]) -> tuple[list[str], list[str], list[str], list[str]]:
    metrics = sorted(
        {
            column["name"]
            for item in items
            for column in item["profile"]["columns"]
            if column["inferred_role"] == "metric"
        }
    )
    times = sorted(
        {
            column["name"]
            for item in items
            for column in item["profile"]["columns"]
            if column["inferred_role"] == "time"
        }
    )
    dimensions = sorted(
        {
            column["name"]
            for item in items
            for column in item["profile"]["columns"]
            if column["inferred_role"] == "dimension"
        }
    )
    identifiers = sorted(
        {
            column["name"]
            for item in items
            for column in item["profile"]["columns"]
            if column["inferred_role"] == "identifier"
        }
    )
    return metrics, times, dimensions, identifiers


def business_field_label(name: str, role: str) -> str:
    normalized = normalize_field_name(name)
    lowered = name.lower()
    for label, hints in BUSINESS_FIELD_HINTS:
        if any(normalize_field_name(hint) in normalized or hint in lowered for hint in hints):
            return label
    role_suffix = {
        "metric": "指标",
        "dimension": "业务对象",
        "time": "时间",
        "identifier": "标识",
    }.get(role, "字段")
    return f"「{name}」{role_suffix}"


def unique_fields(*groups: list[str]) -> list[str]:
    return list(dict.fromkeys(field for group in groups for field in group if field))


def build_capability_summary(items: list[dict[str, Any]], similar_fields: list[dict[str, Any]]) -> dict[str, Any]:
    metrics, times, dimensions, identifiers = discovery_fields(items)
    scope = [
        {"file": item["dataset_path"], "sheet": item["sheet_name"]}
        for item in items
        if item["profile"]["row_count"] > 0
    ]
    date_ranges = [
        column["date_summary"]
        for item in items
        for column in item["profile"]["columns"]
        if column.get("date_summary")
    ]
    return {
        "scope": scope,
        "file_count": len({item["dataset_path"] for item in items}),
        "sheet_count": len(items),
        "row_count": sum(item["profile"]["row_count"] for item in items),
        "metric_fields": metrics,
        "time_fields": times,
        "dimension_fields": dimensions,
        "identifier_fields": identifiers,
        "time_coverage": {
            "min": min((item["min"] for item in date_ranges), default=None),
            "max": max((item["max"] for item in date_ranges), default=None),
        },
        "cross_source_mapping_candidates": similar_fields,
        "reliable_capabilities": [
            capability
            for condition, capability in (
                (bool(metrics), "指标规模与分布"),
                (bool(metrics and times), "时间趋势与波动"),
                (bool(metrics and dimensions), "结构、排名与贡献"),
                (bool(identifiers), "对象级明细追溯"),
            )
            if condition
        ],
        "limited_capabilities": [
            capability
            for condition, capability in (
                (bool(metrics and not times), "趋势分析：缺少明确时间字段"),
                (bool(metrics and not dimensions), "结构归因：缺少可用分析维度"),
                (bool(dimensions and not metrics), "业务结果分析：缺少明确指标"),
                (len(scope) > 1 and not identifiers, "跨表关联：缺少明确关联键"),
            )
            if condition
        ],
        "unsupported_capabilities": [
            "因果结论需要实验、外部事件或业务机制证据，当前结构化数据只能提供诊断线索",
            "预测与优化需要足够时间跨度、稳定口径和独立验证样本",
        ],
    }


def goal_candidate(
    *,
    candidate_id: str,
    category: str,
    title: str,
    question: str,
    decision_value: str,
    scope: list[dict[str, Any]],
    required_fields: list[str],
    supporting_fields: list[str],
    methods: list[str],
    outputs: list[str],
    feasibility: str,
    confidence: str,
    missing_data: list[str] | None = None,
    unsupported: list[str] | None = None,
    cleaning: str = "medium",
    business_metric: str = "",
    business_object: str = "",
) -> dict[str, Any]:
    data_basis = unique_fields(required_fields, supporting_fields)
    return {
        "id": candidate_id,
        "category": category,
        "goal_type": "business_analysis",
        "title": title,
        "business_question": question,
        "questions": [question],
        "decision_value": decision_value,
        "business_metric": business_metric,
        "business_object": business_object,
        "data_basis": data_basis,
        "supported_conclusions": outputs,
        "unsupported_conclusions": unsupported or ["不能仅凭当前数据证明因果关系"],
        "required_scope": scope,
        "required_data": {
            "scope": scope,
            "required_fields": required_fields,
            "supporting_fields": supporting_fields,
        },
        "available_dimensions": supporting_fields,
        "recommended_methods": methods,
        "expected_outputs": outputs,
        "data_quality_risks": [],
        "missing_data": missing_data or [],
        "feasibility": feasibility,
        "confidence": confidence,
        "estimated_cleaning": cleaning,
    }


def build_goal_catalog(items: list[dict[str, Any]], capability: dict[str, Any]) -> dict[str, Any]:
    metrics, times, dimensions, identifiers = discovery_fields(items)
    scope = capability["scope"]
    metric = metrics[0] if metrics else "核心指标"
    time_field = times[0] if times else "时间字段"
    dimension = dimensions[0] if dimensions else "业务维度"
    metric_label = business_field_label(metric, "metric")
    time_label = business_field_label(time_field, "time")
    dimension_label = business_field_label(dimension, "dimension")
    categories = []

    if metrics:
        trend_ready = bool(times)
        categories.append(
            {
                "id": "trend-change",
                "title": "趋势与变化",
                "why_available": f"识别到指标「{metric}」" + (f"和时间字段「{time_field}」" if trend_ready else "，但缺少明确时间字段"),
                "candidates": [
                    goal_candidate(
                        candidate_id="trend-overview",
                        category="趋势与变化",
                        title=(
                            f"{metric_label}是在增长还是下滑，关键变化发生在哪里？"
                            if trend_ready
                            else f"当前{metric_label}处于什么水平，现有数据能否判断变化？"
                        ),
                        question=f"当前{metric_label}表现如何，整体方向和关键变化阶段分别是什么",
                        decision_value="判断整体表现和变化方向",
                        scope=scope,
                        required_fields=[metric] + times[:1],
                        supporting_fields=dimensions[:3],
                        methods=["描述统计", "时间趋势"],
                        outputs=["规模基线", "趋势方向", "关键变化阶段"],
                        feasibility="ready" if trend_ready else "limited",
                        confidence="high" if trend_ready else "medium",
                        missing_data=[] if trend_ready else ["明确时间字段"],
                        business_metric=metric_label,
                    ),
                    goal_candidate(
                        candidate_id="trend-period-comparison",
                        category="趋势与变化",
                        title=f"哪些时间段的{metric_label}最好或最差，差距有多大？",
                        question=f"按{time_label}比较，{metric_label}在哪些期间改善或恶化，变化幅度有多大",
                        decision_value="支持周期复盘和阶段比较",
                        scope=scope,
                        required_fields=[metric] + times[:1],
                        supporting_fields=dimensions[:3],
                        methods=["同比/环比", "期间对比"],
                        outputs=["期间差异", "变化幅度", "重点期间"],
                        feasibility="ready" if trend_ready else "limited",
                        confidence="high" if trend_ready else "low",
                        missing_data=[] if trend_ready else ["明确时间字段和比较周期"],
                        business_metric=metric_label,
                    ),
                    goal_candidate(
                        candidate_id="trend-forecast-readiness",
                        category="趋势与变化",
                        title=f"当前数据能否支持持续监控{metric_label}并提前发现变化？",
                        question=f"现有时间跨度和数据稳定性是否足以为{metric_label}建立监控或预测基线",
                        decision_value="判断能否进入预测和目标管理",
                        scope=scope,
                        required_fields=[metric] + times[:1],
                        supporting_fields=dimensions[:2],
                        methods=["时间覆盖评估", "稳定性检查"],
                        outputs=["监控指标建议", "预测可行性", "待补数据清单"],
                        feasibility="limited",
                        confidence="medium" if trend_ready else "low",
                        missing_data=["足够时间跨度和验证样本"],
                        unsupported=["未验证前不输出正式预测值"],
                        business_metric=metric_label,
                    ),
                ],
            }
        )

    if metrics or dimensions:
        structure_ready = bool(metrics and dimensions)
        categories.append(
            {
                "id": "structure-contribution",
                "title": "结构与贡献",
                "why_available": f"识别到指标「{metric}」和维度「{dimension}」" if structure_ready else "当前只有指标或维度的一侧，部分目标受限",
                "candidates": [
                    goal_candidate(
                        candidate_id="structure-ranking",
                        category="结构与贡献",
                        title=f"哪些{dimension_label}贡献了大部分{metric_label}，应该重点关注谁？",
                        question=f"各{dimension_label}分别贡献了多少{metric_label}，重点对象和集中程度如何",
                        decision_value="识别重点对象和资源集中方向",
                        scope=scope,
                        required_fields=metrics[:1] + dimensions[:1],
                        supporting_fields=times[:1] + dimensions[1:4],
                        methods=["排名分析", "贡献度分析"],
                        outputs=["贡献排名", "集中度", "重点对象"],
                        feasibility="ready" if structure_ready else "limited",
                        confidence="high" if structure_ready else "low",
                        missing_data=[] if structure_ready else ["同时具备指标和业务维度"],
                        business_metric=metric_label,
                        business_object=dimension_label,
                    ),
                    goal_candidate(
                        candidate_id="structure-mix",
                        category="结构与贡献",
                        title=f"{metric_label}主要来自哪些{dimension_label}，是否存在过度集中？",
                        question=f"{metric_label}在不同{dimension_label}之间如何分布，业务结构是否均衡或正在变化",
                        decision_value="理解业务组合和结构风险",
                        scope=scope,
                        required_fields=metrics[:1] + dimensions[:1],
                        supporting_fields=times[:1] + dimensions[1:4],
                        methods=["结构占比", "交叉分析"],
                        outputs=["结构占比", "组合特征", "结构变化"],
                        feasibility="ready" if structure_ready else "limited",
                        confidence="high" if structure_ready else "low",
                        missing_data=[] if structure_ready else ["同时具备指标和业务维度"],
                        business_metric=metric_label,
                        business_object=dimension_label,
                    ),
                    goal_candidate(
                        candidate_id="structure-segmentation",
                        category="结构与贡献",
                        title=f"哪些{dimension_label}属于重点、成长或低表现对象？",
                        question=f"依据{metric_label}和现有业务特征，如何划分{dimension_label}并确定跟进优先级",
                        decision_value="支持差异化管理和后续行动",
                        scope=scope,
                        required_fields=metrics[:1] + dimensions[:1],
                        supporting_fields=dimensions[1:5] + times[:1],
                        methods=["分层规则", "分群画像"],
                        outputs=["分层规则", "对象画像", "重点名单"],
                        feasibility="ready" if structure_ready else "limited",
                        confidence="medium",
                        missing_data=[] if structure_ready else ["稳定指标和对象维度"],
                        business_metric=metric_label,
                        business_object=dimension_label,
                    ),
                ],
            }
        )

    if metrics:
        categories.append(
            {
                "id": "anomaly-risk",
                "title": "异常与风险",
                "why_available": f"数值指标「{metric}」支持分布和异常检测",
                "candidates": [
                    goal_candidate(
                        candidate_id="anomaly-values",
                        category="异常与风险",
                        title=f"哪些{metric_label}记录明显异常，需要优先核查？",
                        question=f"哪些记录的{metric_label}明显偏离常见范围，异常程度和来源分别是什么",
                        decision_value="定位需要核查的异常样本",
                        scope=scope,
                        required_fields=[metric],
                        supporting_fields=dimensions[:4] + times[:1],
                        methods=["分布分析", "IQR异常检测"],
                        outputs=["异常名单", "异常程度", "来源追溯"],
                        feasibility="ready",
                        confidence="medium",
                        cleaning="light",
                        business_metric=metric_label,
                    ),
                    goal_candidate(
                        candidate_id="anomaly-periods",
                        category="异常与风险",
                        title=f"{metric_label}异常主要集中在哪些时间或{dimension_label}？",
                        question=f"异常{metric_label}集中在哪些时间或{dimension_label}，应优先排查哪里",
                        decision_value="缩小问题排查范围",
                        scope=scope,
                        required_fields=[metric],
                        supporting_fields=times[:1] + dimensions[:4],
                        methods=["异常聚合", "切片对比"],
                        outputs=["异常集中区", "异常贡献对象", "核查优先级"],
                        feasibility="ready" if times or dimensions else "limited",
                        confidence="medium",
                        missing_data=[] if times or dimensions else ["时间或业务维度"],
                        business_metric=metric_label,
                        business_object=dimension_label,
                    ),
                    goal_candidate(
                        candidate_id="anomaly-monitoring",
                        category="异常与风险",
                        title=f"{metric_label}达到什么情况时应该触发预警？",
                        question=f"应如何持续监控{metric_label}，并按影响程度设置分级预警和处置规则",
                        decision_value="形成可复用的风险监控机制",
                        scope=scope,
                        required_fields=[metric],
                        supporting_fields=times[:1] + dimensions[:3],
                        methods=["阈值评估", "监控规则设计"],
                        outputs=["监控口径", "预警阈值建议", "处置分级"],
                        feasibility="limited",
                        confidence="medium",
                        missing_data=["业务容忍范围和历史处置结果"],
                        business_metric=metric_label,
                    ),
                ],
            }
        )

    governance_fields = (times + metrics + dimensions + identifiers)[:12]
    categories.append(
        {
            "id": "data-governance",
            "title": "数据治理与可用性",
            "why_available": "数据侦察已识别表头、字段角色、跨来源差异和基础质量",
            "candidates": [
                {
                    **goal_candidate(
                        candidate_id="governance-standardize",
                        category="数据治理与可用性",
                        title="能否整理出一份可直接分析、后续还能复用的标准明细？",
                        question="如何统一目标相关字段、粒度和来源血缘，供后续分析使用",
                        decision_value="降低重复整理成本",
                        scope=scope,
                        required_fields=[],
                        supporting_fields=governance_fields,
                        methods=["字段标准化", "结构归一"],
                        outputs=["标准明细文件", "字段映射", "清洗审计"],
                        feasibility="ready" if scope else "blocked",
                        confidence="high" if scope else "low",
                    ),
                    "goal_type": "operational_cleaning",
                },
                {
                    **goal_candidate(
                        candidate_id="governance-quality",
                        category="数据治理与可用性",
                        title="哪些数据问题会影响业务结论，应该优先处理什么？",
                        question="当前数据问题会影响哪些分析结论，应如何处置",
                        decision_value="明确数据可用边界和修复优先级",
                        scope=scope,
                        required_fields=[],
                        supporting_fields=governance_fields,
                        methods=["质量画像", "结论影响分级"],
                        outputs=["质量问题清单", "影响级别", "处置建议"],
                        feasibility="ready" if scope else "blocked",
                        confidence="high" if scope else "low",
                        cleaning="light",
                    ),
                    "goal_type": "operational_cleaning",
                },
                {
                    **goal_candidate(
                        candidate_id="governance-integration",
                        category="数据治理与可用性",
                        title="这些文件能否可靠合并，哪些字段需要先统一？",
                        question="不同文件和Sheet能否可靠合并或关联，需要哪些标准化规则",
                        decision_value="建立后续多目标分析的数据底座",
                        scope=scope,
                        required_fields=[],
                        supporting_fields=governance_fields,
                        methods=["结构对比", "关联键评估"],
                        outputs=["合并方案", "关联风险", "待补字段"],
                        feasibility="ready" if len(scope) > 1 else "limited",
                        confidence="medium",
                        missing_data=[] if len(scope) > 1 else ["更多需要整合的数据来源"],
                    ),
                    "goal_type": "operational_cleaning",
                },
            ],
        }
    )
    for category in categories:
        for candidate in category["candidates"]:
            candidate["time_coverage"] = capability.get("time_coverage", {})
            candidate["data_scope_summary"] = {
                "file_count": capability.get("file_count", 0),
                "sheet_count": capability.get("sheet_count", 0),
                "row_count": capability.get("row_count", 0),
            }
    return {
        "schema_version": "1.0",
        "generated_at": datetime.now().astimezone().isoformat(),
        "categories": categories,
        "custom_goal_available": True,
        "candidate_count": sum(len(category["candidates"]) for category in categories),
    }


def build_discovery(paths: list[Path], sheet_name: str | None, all_sheets: bool) -> dict[str, Any]:
    from clean_tabular_data import build_item, load_sources

    items = []
    field_index: dict[str, list[dict[str, Any]]] = {}
    sources = load_sources(paths, sheet_name)
    for source in sources:
        structure = build_item(source)
        if structure.get("detected"):
            headers = structure["headers"]
            rows = [
                {
                    header: source["rows"][row_number - 1][index]
                    if index < len(source["rows"][row_number - 1])
                    else None
                    for index, header in enumerate(headers)
                }
                for row_number in structure.get("included_rows", [])
            ]
        else:
            rows = []
        profile = build_profile(source["path"], rows)
        item = {
                "dataset_path": str(source["path"]),
                "sheet_name": source["sheet"],
                "structure": {
                    key: structure.get(key)
                    for key in (
                        "detected",
                        "confidence",
                        "header_row",
                        "header_rows",
                        "data_start_row",
                        "data_end_row",
                        "included_row_count",
                        "exclusion_candidate_counts",
                        "evidence",
                        "blocking_reasons",
                    )
                },
                "profile": profile,
            }
        items.append(item)
        for column in profile["columns"]:
            normalized = normalize_field_name(column["name"])
            field_index.setdefault(normalized, []).append(
                {
                    "file": str(source["path"]),
                    "sheet": item["sheet_name"],
                    "field": column["name"],
                    "role": column["inferred_role"],
                }
            )
    similar_fields = [
        {"normalized_name": key, "occurrences": occurrences}
        for key, occurrences in field_index.items()
        if key and len({entry["field"] for entry in occurrences}) > 1
    ]
    capability = build_capability_summary(items, similar_fields)
    goal_catalog = build_goal_catalog(items, capability)
    payload = {
        "schema_version": "2.0",
        "generated_at": datetime.now().astimezone().isoformat(),
        "mode": "discovery",
        "items": items,
        "similar_fields": similar_fields,
        "capability_summary": capability,
        "goal_catalog": goal_catalog,
    }
    payload["goal_candidates"] = [
        candidate
        for category in goal_catalog["categories"]
        for candidate in category["candidates"]
    ]
    payload["discovery_id"] = contract_id(payload)
    return payload


def main() -> None:
    args = parse_args()
    dataset_paths = [Path(item).expanduser().resolve() for item in args.dataset]
    if len(dataset_paths) == 1 and not args.all_sheets:
        dataset_path = dataset_paths[0]
        rows = load_rows(dataset_path, args.sheet)
        profile = build_profile(dataset_path, rows)
    else:
        profile = build_discovery(dataset_paths, args.sheet, args.all_sheets)
    payload = json.dumps(profile, ensure_ascii=False, indent=2)
    if args.output:
        output_path = Path(args.output)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        output_path.write_text(payload + "\n", encoding="utf-8")
    print(payload)


if __name__ == "__main__":
    main()
