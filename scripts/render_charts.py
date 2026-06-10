#!/usr/bin/env python3
"""Render confirmed chart candidates from a targeted analysis dataset."""

from __future__ import annotations

import argparse
import csv
import json
import math
import re
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path
from typing import Any, Callable

from openpyxl import load_workbook
from PIL import Image, ImageDraw, ImageFont


WIDTH = 1400
HEIGHT = 820
PLOT = (150, 120, 1260, 680)
COLORS = ["#175CD3", "#12B76A", "#F79009", "#D92D20", "#7A5AF8", "#06AED4", "#EE46BC", "#667085"]
GRID = "#D0D5DD"
INK = "#17202A"
MUTED = "#667085"
FONT_CANDIDATES = [
    Path("/System/Library/Fonts/Hiragino Sans GB.ttc"),
    Path("/System/Library/Fonts/STHeiti Medium.ttc"),
    Path("/Library/Fonts/Arial Unicode.ttf"),
    Path("/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc"),
]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--dataset", required=True)
    parser.add_argument("--decision", required=True)
    parser.add_argument("--output-dir", required=True)
    parser.add_argument("--run-file", help="Optional analysis run JSON to update")
    parser.add_argument("--dpi", type=int, default=160)
    return parser.parse_args()


def read_json(path: Path) -> dict[str, Any]:
    return json.loads(path.read_text(encoding="utf-8"))


def write_json(path: Path, payload: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def font(size: int) -> ImageFont.FreeTypeFont | ImageFont.ImageFont:
    for candidate in FONT_CANDIDATES:
        if candidate.exists():
            return ImageFont.truetype(str(candidate), size=size)
    return ImageFont.load_default()


FONTS = {
    "title": font(34),
    "label": font(22),
    "small": font(18),
    "tiny": font(15),
}


def load_dataset(path: Path, sheet: str) -> list[dict[str, Any]]:
    suffix = path.suffix.lower()
    if suffix in {".xlsx", ".xlsm"}:
        workbook = load_workbook(path, read_only=True, data_only=True)
        worksheet = workbook[sheet] if sheet and sheet != "未指定" else workbook.active
        rows = worksheet.iter_rows(values_only=True)
        headers = [str(value).strip() if value is not None else "" for value in next(rows)]
        return [dict(zip(headers, values)) for values in rows]
    if suffix == ".csv":
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            return list(csv.DictReader(handle))
    raise ValueError(f"Unsupported dataset format: {suffix}")


def safe_filename(candidate_id: str) -> str:
    cleaned = re.sub(r"[^A-Za-z0-9._-]+", "-", candidate_id).strip("-")
    return cleaned or "chart"


def number(value: Any) -> float | None:
    if value is None or value == "":
        return None
    try:
        parsed = float(str(value).replace(",", "").replace("%", "").strip())
    except (TypeError, ValueError):
        return None
    return parsed if math.isfinite(parsed) else None


def time_key(value: Any) -> tuple[int, Any]:
    if isinstance(value, (datetime, date)):
        return (0, value.isoformat())
    text = str(value).strip()
    normalized = text.replace("/", "-").replace("年", "-").replace("月", "-").replace("日", "")
    try:
        return (0, datetime.fromisoformat(normalized).isoformat())
    except ValueError:
        return (1, text)


def require_columns(rows: list[dict[str, Any]], fields: list[str]) -> None:
    available = set(rows[0]) if rows else set()
    missing = [field for field in fields if field not in available]
    if missing:
        raise ValueError(f"Missing fields: {', '.join(missing)}")


def text_width(draw: ImageDraw.ImageDraw, text: str, selected_font: Any) -> int:
    box = draw.textbbox((0, 0), text, font=selected_font)
    return box[2] - box[0]


def shorten(draw: ImageDraw.ImageDraw, text: Any, max_width: int, selected_font: Any) -> str:
    value = str(text)
    if text_width(draw, value, selected_font) <= max_width:
        return value
    while value and text_width(draw, value + "…", selected_font) > max_width:
        value = value[:-1]
    return value + "…"


def base_image(title: str) -> tuple[Image.Image, ImageDraw.ImageDraw]:
    image = Image.new("RGB", (WIDTH, HEIGHT), "white")
    draw = ImageDraw.Draw(image)
    draw.text((70, 42), title, fill=INK, font=FONTS["title"])
    draw.line((70, 92, WIDTH - 70, 92), fill=GRID, width=2)
    return image, draw


def draw_axes(draw: ImageDraw.ImageDraw, x_label: str = "", y_label: str = "") -> None:
    left, top, right, bottom = PLOT
    draw.line((left, top, left, bottom), fill=INK, width=2)
    draw.line((left, bottom, right, bottom), fill=INK, width=2)
    if x_label:
        draw.text(((left + right) // 2, bottom + 78), x_label, fill=MUTED, font=FONTS["label"], anchor="mm")
    if y_label:
        draw.text((30, (top + bottom) // 2), y_label, fill=MUTED, font=FONTS["label"], anchor="lm")


def numeric_ticks(draw: ImageDraw.ImageDraw, minimum: float, maximum: float, horizontal: bool = False) -> None:
    left, top, right, bottom = PLOT
    if math.isclose(minimum, maximum):
        maximum = minimum + 1
    for index in range(6):
        ratio = index / 5
        value = minimum + (maximum - minimum) * ratio
        if horizontal:
            x = left + int((right - left) * ratio)
            draw.line((x, top, x, bottom), fill=GRID, width=1)
            draw.text((x, bottom + 12), f"{value:,.1f}", fill=MUTED, font=FONTS["tiny"], anchor="ma")
        else:
            y = bottom - int((bottom - top) * ratio)
            draw.line((left, y, right, y), fill=GRID, width=1)
            draw.text((left - 12, y), f"{value:,.1f}", fill=MUTED, font=FONTS["tiny"], anchor="rm")


def grouped_metric(rows: list[dict[str, Any]], dimension: str, metric: str) -> list[tuple[str, float]]:
    grouped: dict[str, float] = defaultdict(float)
    for row in rows:
        value = number(row.get(metric))
        label = row.get(dimension)
        if value is None or label in {None, ""}:
            continue
        grouped[str(label)] += value
    if not grouped:
        raise ValueError("No valid dimension and metric rows")
    return sorted(grouped.items(), key=lambda item: item[1], reverse=True)


def render_trend(draw: ImageDraw.ImageDraw, rows: list[dict[str, Any]], fields: list[str]) -> None:
    time_field, metric = fields[:2]
    grouped: dict[str, float] = defaultdict(float)
    for row in rows:
        value = number(row.get(metric))
        label = row.get(time_field)
        if value is not None and label not in {None, ""}:
            grouped[str(label)] += value
    points = sorted(grouped.items(), key=lambda item: time_key(item[0]))
    if not points:
        raise ValueError("No valid time and metric rows")
    values = [item[1] for item in points]
    maximum = max(values)
    minimum = min(0.0, min(values))
    draw_axes(draw, time_field, metric)
    numeric_ticks(draw, minimum, maximum)
    left, top, right, bottom = PLOT
    coordinates = []
    for index, (label, value) in enumerate(points):
        x = left + int((right - left) * (index / max(1, len(points) - 1)))
        y = bottom - int((bottom - top) * ((value - minimum) / max(1e-9, maximum - minimum)))
        coordinates.append((x, y))
        if index % max(1, math.ceil(len(points) / 8)) == 0:
            draw.text((x, bottom + 12), shorten(draw, label, 110, FONTS["tiny"]), fill=MUTED, font=FONTS["tiny"], anchor="ma")
    if len(coordinates) > 1:
        draw.line(coordinates, fill=COLORS[0], width=5)
    for x, y in coordinates:
        draw.ellipse((x - 6, y - 6, x + 6, y + 6), fill=COLORS[0])


def render_comparison(draw: ImageDraw.ImageDraw, rows: list[dict[str, Any]], fields: list[str]) -> None:
    dimension, metric = fields[:2]
    grouped = grouped_metric(rows, dimension, metric)[:15]
    maximum = max(value for _, value in grouped)
    draw_axes(draw, metric, dimension)
    numeric_ticks(draw, 0, maximum, horizontal=True)
    left, top, right, bottom = PLOT
    slot = (bottom - top) / len(grouped)
    for index, (label, value) in enumerate(reversed(grouped)):
        y1 = top + index * slot + 5
        y2 = top + (index + 1) * slot - 5
        x2 = left + int((right - left) * value / max(1e-9, maximum))
        draw.rectangle((left, y1, x2, y2), fill=COLORS[0])
        draw.text((left - 12, (y1 + y2) / 2), shorten(draw, label, 125, FONTS["tiny"]), fill=MUTED, font=FONTS["tiny"], anchor="rm")
        draw.text((x2 + 8, (y1 + y2) / 2), f"{value:,.1f}", fill=INK, font=FONTS["tiny"], anchor="lm")


def render_composition(draw: ImageDraw.ImageDraw, rows: list[dict[str, Any]], fields: list[str]) -> None:
    if len(fields) < 3:
        dimension, metric = fields[:2]
        grouped = grouped_metric(rows, dimension, metric)[:8]
        total = sum(value for _, value in grouped)
        if math.isclose(total, 0):
            raise ValueError("Composition total is zero")
        left, top, right, bottom = PLOT
        y1, y2 = (top + bottom) // 2 - 55, (top + bottom) // 2 + 55
        cursor = left
        for index, (label, value) in enumerate(grouped):
            width = int((right - left) * value / total)
            draw.rectangle((cursor, y1, cursor + width, y2), fill=COLORS[index % len(COLORS)])
            draw.text((cursor + width / 2, y2 + 18), shorten(draw, label, max(50, width), FONTS["tiny"]), fill=MUTED, font=FONTS["tiny"], anchor="ma")
            cursor += width
        draw.text(((left + right) / 2, bottom + 62), metric, fill=MUTED, font=FONTS["label"], anchor="mm")
        return

    time_field, dimension, metric = fields[:3]
    top_dimensions = [label for label, _ in grouped_metric(rows, dimension, metric)[:8]]
    grouped: dict[str, dict[str, float]] = defaultdict(lambda: defaultdict(float))
    for row in rows:
        value = number(row.get(metric))
        time_value = row.get(time_field)
        dimension_value = str(row.get(dimension))
        if value is not None and time_value not in {None, ""} and dimension_value in top_dimensions:
            grouped[str(time_value)][dimension_value] += value
    periods = sorted(grouped, key=time_key)
    if not periods:
        raise ValueError("No valid composition rows")
    totals = [sum(grouped[period].values()) for period in periods]
    maximum = max(totals)
    draw_axes(draw, time_field, metric)
    numeric_ticks(draw, 0, maximum)
    left, top, right, bottom = PLOT
    slot = (right - left) / len(periods)
    bar_width = max(8, int(slot * 0.68))
    for index, period in enumerate(periods):
        x1 = left + index * slot + (slot - bar_width) / 2
        cursor = bottom
        for color_index, label in enumerate(top_dimensions):
            value = grouped[period].get(label, 0)
            height = (bottom - top) * value / max(1e-9, maximum)
            draw.rectangle((x1, cursor - height, x1 + bar_width, cursor), fill=COLORS[color_index % len(COLORS)])
            cursor -= height
        if index % max(1, math.ceil(len(periods) / 8)) == 0:
            draw.text((x1 + bar_width / 2, bottom + 12), shorten(draw, period, 100, FONTS["tiny"]), fill=MUTED, font=FONTS["tiny"], anchor="ma")
    legend_x = right + 15
    for index, label in enumerate(top_dimensions):
        y = top + index * 34
        draw.rectangle((legend_x, y, legend_x + 18, y + 18), fill=COLORS[index % len(COLORS)])
        draw.text((legend_x + 25, y + 9), shorten(draw, label, 100, FONTS["tiny"]), fill=MUTED, font=FONTS["tiny"], anchor="lm")


def render_pareto(draw: ImageDraw.ImageDraw, rows: list[dict[str, Any]], fields: list[str]) -> None:
    dimension, metric = fields[:2]
    grouped = grouped_metric(rows, dimension, metric)[:15]
    total = sum(value for _, value in grouped)
    if math.isclose(total, 0):
        raise ValueError("Pareto total is zero")
    maximum = max(value for _, value in grouped)
    draw_axes(draw, dimension, metric)
    numeric_ticks(draw, 0, maximum)
    left, top, right, bottom = PLOT
    slot = (right - left) / len(grouped)
    cumulative = 0.0
    points = []
    for index, (label, value) in enumerate(grouped):
        x1 = left + index * slot + slot * 0.15
        x2 = left + (index + 1) * slot - slot * 0.15
        y = bottom - (bottom - top) * value / max(1e-9, maximum)
        draw.rectangle((x1, y, x2, bottom), fill=COLORS[0])
        draw.text(((x1 + x2) / 2, bottom + 12), shorten(draw, label, int(slot), FONTS["tiny"]), fill=MUTED, font=FONTS["tiny"], anchor="ma")
        cumulative += value
        points.append(((x1 + x2) / 2, bottom - (bottom - top) * cumulative / total))
    if len(points) > 1:
        draw.line(points, fill=COLORS[3], width=4)
    for x, y in points:
        draw.ellipse((x - 5, y - 5, x + 5, y + 5), fill=COLORS[3])


def render_distribution(draw: ImageDraw.ImageDraw, rows: list[dict[str, Any]], fields: list[str]) -> None:
    metric = fields[0]
    values = sorted(value for row in rows if (value := number(row.get(metric))) is not None)
    if not values:
        raise ValueError("No valid metric rows")

    def percentile(ratio: float) -> float:
        position = (len(values) - 1) * ratio
        lower = math.floor(position)
        upper = math.ceil(position)
        if lower == upper:
            return values[lower]
        return values[lower] * (upper - position) + values[upper] * (position - lower)

    low, q1, median, q3, high = values[0], percentile(0.25), percentile(0.5), percentile(0.75), values[-1]
    draw_axes(draw, metric, "")
    numeric_ticks(draw, low, high, horizontal=True)
    left, top, right, bottom = PLOT

    def x(value: float) -> float:
        return left + (right - left) * (value - low) / max(1e-9, high - low)

    center = (top + bottom) / 2
    draw.line((x(low), center, x(high), center), fill=INK, width=4)
    draw.rectangle((x(q1), center - 65, x(q3), center + 65), fill="#84ADFF", outline=COLORS[0], width=3)
    draw.line((x(median), center - 65, x(median), center + 65), fill=COLORS[3], width=4)
    for value in (low, high):
        draw.line((x(value), center - 35, x(value), center + 35), fill=INK, width=3)


def render_relationship(draw: ImageDraw.ImageDraw, rows: list[dict[str, Any]], fields: list[str]) -> None:
    first, second = fields[:2]
    pairs = []
    for row in rows:
        x_value, y_value = number(row.get(first)), number(row.get(second))
        if x_value is not None and y_value is not None:
            pairs.append((x_value, y_value))
    if not pairs:
        raise ValueError("No valid paired metric rows")
    x_min, x_max = min(item[0] for item in pairs), max(item[0] for item in pairs)
    y_min, y_max = min(item[1] for item in pairs), max(item[1] for item in pairs)
    draw_axes(draw, first, second)
    numeric_ticks(draw, y_min, y_max)
    left, top, right, bottom = PLOT
    for x_value, y_value in pairs:
        x = left + (right - left) * (x_value - x_min) / max(1e-9, x_max - x_min)
        y = bottom - (bottom - top) * (y_value - y_min) / max(1e-9, y_max - y_min)
        draw.ellipse((x - 7, y - 7, x + 7, y + 7), fill=COLORS[0])


RENDERERS: dict[str, Callable[[ImageDraw.ImageDraw, list[dict[str, Any]], list[str]], None]] = {
    "trend": render_trend,
    "comparison": render_comparison,
    "composition": render_composition,
    "pareto": render_pareto,
    "distribution": render_distribution,
    "relationship": render_relationship,
}


def update_run(run_file: Path, decision: dict[str, Any], result: dict[str, Any]) -> None:
    run = read_json(run_file) if run_file.exists() else {}
    run["chart_decision"] = decision
    run["chart_files"] = [item["path"] for item in result["generated"]]
    run["chart_failures"] = result["failed"]
    run["saved_at"] = datetime.now().astimezone().isoformat()
    write_json(run_file, run)


def main() -> None:
    args = parse_args()
    dataset = Path(args.dataset).expanduser().resolve()
    decision_path = Path(args.decision).expanduser().resolve()
    output_dir = Path(args.output_dir).expanduser().resolve()
    decision = read_json(decision_path)
    if decision.get("status") != "confirmed":
        raise SystemExit("Chart decision must be confirmed before rendering")

    output_dir.mkdir(parents=True, exist_ok=True)
    selected = decision.get("selected_charts", [])
    rows = load_dataset(dataset, decision.get("sheet", "未指定")) if selected else []
    generated: list[dict[str, Any]] = []
    failed: list[dict[str, Any]] = []

    for candidate in selected:
        candidate_id = candidate.get("id", "")
        fields = candidate.get("fields", [])
        try:
            renderer = RENDERERS[candidate_id]
            require_columns(rows, fields)
            image, draw = base_image(candidate.get("title", candidate_id))
            renderer(draw, rows, fields)
            output_path = output_dir / f"{safe_filename(candidate_id)}.png"
            image.save(output_path, format="PNG", dpi=(args.dpi, args.dpi))
            generated.append(
                {
                    "id": candidate_id,
                    "title": candidate.get("title", candidate_id),
                    "chart_type": candidate.get("chart_type", ""),
                    "path": str(output_path),
                    "markdown": f"![{candidate.get('title', candidate_id)}]({output_path})",
                }
            )
        except Exception as exc:
            failed.append(
                {
                    "id": candidate_id,
                    "title": candidate.get("title", candidate_id),
                    "error": str(exc),
                }
            )

    result = {
        "generated_at": datetime.now().astimezone().isoformat(),
        "decision_path": str(decision_path),
        "output_dir": str(output_dir),
        "generated": generated,
        "failed": failed,
        "chart_files": [item["path"] for item in generated],
    }
    if args.run_file:
        update_run(Path(args.run_file).expanduser().resolve(), decision, result)
    print(json.dumps(result, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
