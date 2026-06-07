#!/usr/bin/env python3
"""Clean Excel/CSV exports by detecting header and data regions and writing a new XLSX file."""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import math
import re
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

try:
    from openpyxl import Workbook, load_workbook
except ImportError:  # pragma: no cover - optional dependency
    Workbook = None
    load_workbook = None


FIELD_KEYWORDS = (
    "订单",
    "单号",
    "编号",
    "日期",
    "时间",
    "金额",
    "费用",
    "成本",
    "收入",
    "税",
    "币种",
    "姓名",
    "人员",
    "员工",
    "部门",
    "公司",
    "城市",
    "航程",
    "航班",
    "order",
    "date",
    "time",
    "amount",
    "price",
    "cost",
    "tax",
    "name",
    "dept",
    "city",
    "flight",
)
SUMMARY_LABELS = ("合计", "总计", "小计", "汇总")
NOTE_LABELS = ("备注", "说明")
SIGNOFF_LABELS = ("制表", "审核", "复核")
AUDIT_SHEET_NAME = "清洗排除记录"


@dataclass
class Region:
    header_row: int
    header_rows: list[int]
    data_start_row: int | None
    data_end_row: int | None
    included_rows: list[int]
    skipped_rows: list[dict[str, Any]]
    exclusion_candidates: list[dict[str, Any]]
    confidence: str
    evidence: list[str]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("inputs", nargs="+", help="Input .xlsx/.csv files. Excel temporary files starting with .~ are ignored.")
    parser.add_argument("--output", help="Output .xlsx path. Defaults to <input>_清洗后.xlsx for one input.")
    parser.add_argument("--sheet", help="Worksheet name to process. Defaults to all visible worksheets.")
    parser.add_argument("--target-sheet", default="原始数据_清洗后")
    parser.add_argument("--profile-output", help="Write structure profile JSON.")
    parser.add_argument("--rules", help="Use a previously confirmed rules JSON instead of auto-detecting regions.")
    parser.add_argument("--rules-output", help="Write detected cleaning rules JSON.")
    parser.add_argument("--run-output", help="Write dry-run or execution summary JSON.")
    parser.add_argument("--handoff-output", help="Write analysis handoff JSON after successful execution.")
    parser.add_argument("--cleaning-run-id", help="Cleaning run ID recorded in the handoff JSON.")
    merge_mode = parser.add_mutually_exclusive_group()
    merge_mode.add_argument(
        "--merge-same-sheets",
        action="store_true",
        help="Merge same-named worksheets from different files into one output worksheet.",
    )
    merge_mode.add_argument(
        "--keep-separate-sheets",
        action="store_true",
        help="Keep same-named worksheets separate and suffix output worksheet names with source filenames.",
    )
    parser.add_argument("--execute", action="store_true", help="Write the cleaned data to a new XLSX file.")
    parser.add_argument("--no-lineage", action="store_true", help="Do not add source filename/sheet/row columns.")
    return parser.parse_args()


def cell_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and math.isnan(value):
        return ""
    return str(value).strip()


def is_blank_row(row: list[Any]) -> bool:
    return all(cell_text(value) == "" for value in row)


def non_empty_values(row: list[Any]) -> list[Any]:
    return [value for value in row if cell_text(value) != ""]


def parse_number(value: Any) -> float | None:
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        if isinstance(value, float) and math.isnan(value):
            return None
        return float(value)
    text = cell_text(value).replace(",", "")
    if not text:
        return None
    if text.endswith("%"):
        text = text[:-1]
    try:
        return float(text)
    except ValueError:
        return None


def parse_date(value: Any) -> datetime | None:
    if isinstance(value, datetime):
        return value
    text = cell_text(value)
    if not text:
        return None
    if text.isdigit() and len(text) not in {6, 8}:
        return None
    formats = (
        "%Y-%m-%d",
        "%Y/%m/%d",
        "%Y-%m",
        "%Y/%m",
        "%Y%m%d",
        "%Y%m",
        "%m/%d/%Y",
        "%d/%m/%Y",
        "%Y-%m-%d %H:%M:%S",
        "%Y/%m/%d %H:%M:%S",
    )
    for fmt in formats:
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    return None


def trim_trailing_empty(row: list[Any]) -> list[Any]:
    end = len(row)
    while end > 0 and cell_text(row[end - 1]) == "":
        end -= 1
    return row[:end]


def row_width(row: list[Any]) -> int:
    return len(non_empty_values(row))


def short_text_ratio(row: list[Any]) -> float:
    values = non_empty_values(row)
    if not values:
        return 0
    short_texts = 0
    for value in values:
        text = cell_text(value)
        if parse_number(value) is None and parse_date(value) is None and len(text) <= 30:
            short_texts += 1
    return short_texts / len(values)


def keyword_hits(row: list[Any]) -> int:
    lowered = [cell_text(value).lower() for value in non_empty_values(row)]
    return sum(1 for text in lowered for keyword in FIELD_KEYWORDS if keyword in text)


def duplicate_rate(row: list[Any]) -> float:
    values = [cell_text(value) for value in non_empty_values(row)]
    if not values:
        return 1
    return 1 - (len(set(values)) / len(values))


def data_signal_count(row: list[Any]) -> int:
    signals = 0
    for value in non_empty_values(row):
        text = cell_text(value)
        if parse_date(value) is not None:
            signals += 1
        if parse_number(value) is not None:
            signals += 1
        if re.fullmatch(r"\d{8,}", text):
            signals += 1
        if re.search(r"[\u4e00-\u9fffA-Za-z]+[-/][\u4e00-\u9fffA-Za-z]+", text):
            signals += 1
        if re.fullmatch(r"[A-Z]{1,3}\d{2,5}", text):
            signals += 1
    return signals


def row_content_hash(row: list[Any]) -> str:
    payload = json.dumps([cell_text(value) for value in row], ensure_ascii=False, separators=(",", ":"))
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()


def matched_cell(column: int, value: Any, label: str) -> dict[str, Any]:
    return {"column": column, "value": cell_text(value), "label": label}


def exclusion_candidate(row: list[Any], expected_width: int) -> dict[str, Any] | None:
    populated = [(idx + 1, value, cell_text(value)) for idx, value in enumerate(row) if cell_text(value)]
    if not populated:
        return None

    width = len(populated)
    sparse_limit = max(2, math.ceil(expected_width * 0.35))
    is_sparse = width <= sparse_limit
    numeric_count = sum(1 for _, value, _ in populated if parse_number(value) is not None)

    for column, value, text in populated[:2]:
        normalized = re.sub(r"[\s:：_-]+", "", text).lower()
        summary_match = next(
            (
                label
                for label in SUMMARY_LABELS
                if normalized == label or (normalized.endswith(label) and len(normalized) <= 12)
            ),
            None,
        )
        english_summary = normalized in {"subtotal", "total", "grandtotal"}
        if (summary_match or english_summary) and (is_sparse or numeric_count > 0):
            label = summary_match or normalized
            return {
                "category": "summary",
                "reason": "summary_row_candidate",
                "matched_cells": [matched_cell(column, value, label)],
                "evidence": f"前两个非空单元格中出现汇总标签“{text}”，且该行稀疏或包含数值",
                "suggested_action": "exclude",
            }

    first_column, first_value, first_text = populated[0]
    normalized_first = first_text.strip().lower()
    if is_sparse:
        for label in NOTE_LABELS:
            if re.fullmatch(rf"{label}(?:\s*[:：].*)?", normalized_first, flags=re.IGNORECASE):
                return {
                    "category": "note",
                    "reason": "note_row_candidate",
                    "matched_cells": [matched_cell(first_column, first_value, label)],
                    "evidence": f"首个非空单元格为说明标签“{first_text}”，且该行明显稀疏",
                    "suggested_action": "exclude",
                }
        if re.fullmatch(r"note(?:\s*[:：].*)?", normalized_first, flags=re.IGNORECASE):
            return {
                "category": "note",
                "reason": "note_row_candidate",
                "matched_cells": [matched_cell(first_column, first_value, "note")],
                "evidence": f"首个非空单元格为说明标签“{first_text}”，且该行明显稀疏",
                "suggested_action": "exclude",
            }
        for label in SIGNOFF_LABELS:
            if re.fullmatch(rf"{label}(?:人|员)?(?:\s*[:：].*)?", normalized_first):
                return {
                    "category": "signoff",
                    "reason": "signoff_row_candidate",
                    "matched_cells": [matched_cell(first_column, first_value, label)],
                    "evidence": f"首个非空单元格为签字标签“{first_text}”，且该行明显稀疏",
                    "suggested_action": "exclude",
                }

    if width < max(2, expected_width * 0.2):
        return {
            "category": "sparse",
            "reason": "too_sparse_candidate",
            "matched_cells": [],
            "evidence": f"该行仅有 {width} 个非空单元格，显著少于表头宽度 {expected_width}",
            "suggested_action": "exclude",
        }
    return None


def row_similarity(a: list[Any], b: list[Any]) -> float:
    aw = row_width(a)
    bw = row_width(b)
    if max(aw, bw) == 0:
        return 0
    return min(aw, bw) / max(aw, bw)


def looks_like_english_header(row: list[Any], expected_width: int) -> bool:
    values = non_empty_values(row)
    if len(values) < max(3, expected_width * 0.5):
        return False
    ascii_like = 0
    for value in values:
        text = cell_text(value)
        if re.fullmatch(r"[A-Za-z_][A-Za-z0-9_ /.-]{0,40}", text):
            ascii_like += 1
    return ascii_like / len(values) >= 0.65 and data_signal_count(row) <= 1 and short_text_ratio(row) >= 0.8


def looks_like_data_row(row: list[Any], expected_width: int) -> bool:
    if is_blank_row(row) or exclusion_candidate(row, expected_width):
        return False
    values = non_empty_values(row)
    if len(values) < max(3, expected_width * 0.35):
        return False
    if data_signal_count(row) >= 2:
        return True
    return short_text_ratio(row) < 0.75 and len(values) >= max(5, expected_width * 0.5)


def header_score(rows: list[list[Any]], idx: int) -> tuple[float, list[str]]:
    row = rows[idx]
    width = row_width(row)
    score = 0.0
    evidence: list[str] = []
    if width >= 5:
        score += 2
        evidence.append(f"row has {width} non-empty cells")
    ratio = short_text_ratio(row)
    if ratio >= 0.7:
        score += 2
        evidence.append(f"short-text ratio {ratio:.2f}")
    hits = keyword_hits(row)
    if hits:
        score += min(3, hits)
        evidence.append(f"{hits} field keyword hits")
    dup = duplicate_rate(row)
    if dup <= 0.2:
        score += 1
        evidence.append(f"low duplicate rate {dup:.2f}")
    if data_signal_count(row) <= 1:
        score += 1
        evidence.append("row does not look like detail data")
    for lookahead in (1, 2, 3):
        if idx + lookahead >= len(rows):
            continue
        next_row = rows[idx + lookahead]
        if is_blank_row(next_row):
            continue
        similarity = row_similarity(row, next_row)
        if similarity >= 0.55:
            score += 1
            evidence.append(f"next row +{lookahead} has similar width")
            break
    for lookahead in range(1, 5):
        if idx + lookahead < len(rows) and looks_like_data_row(rows[idx + lookahead], width):
            score += 2
            evidence.append(f"detail-like row appears at +{lookahead}")
            break
    return score, evidence


def detect_region(rows: list[list[Any]]) -> Region | None:
    candidates = []
    scan_limit = min(len(rows), 120)
    for idx in range(scan_limit):
        if is_blank_row(rows[idx]):
            continue
        score, evidence = header_score(rows, idx)
        if score >= 5:
            candidates.append((score, idx, evidence))
    if not candidates:
        return None
    candidates.sort(key=lambda item: (-item[0], item[1]))
    score, header_idx, evidence = candidates[0]
    if header_idx > 0:
        previous_row = rows[header_idx - 1]
        current_width = max(row_width(rows[header_idx]), 1)
        previous_width = row_width(previous_row)
        if (
            looks_like_english_header(rows[header_idx], previous_width)
            and previous_width >= 5
            and row_similarity(previous_row, rows[header_idx]) >= 0.55
            and short_text_ratio(previous_row) >= 0.7
            and keyword_hits(previous_row) > 0
        ):
            score, previous_evidence = header_score(rows, header_idx - 1)
            header_idx -= 1
            evidence = previous_evidence + [
                f"next row has {current_width} English/system-like fields and is treated as secondary header"
            ]
    header_width = max(len(trim_trailing_empty(rows[header_idx])), row_width(rows[header_idx]))
    header_rows = [header_idx + 1]
    cursor = header_idx + 1
    if cursor < len(rows) and looks_like_english_header(rows[cursor], header_width):
        header_rows.append(cursor + 1)
        evidence.append("next row looks like English/system header and will be skipped")
        cursor += 1

    data_start_idx: int | None = None
    for idx in range(cursor, len(rows)):
        if not looks_like_data_row(rows[idx], header_width):
            continue
        window = rows[idx : min(len(rows), idx + 5)]
        detail_count = sum(1 for item in window if looks_like_data_row(item, header_width))
        if detail_count >= min(3, len(window)):
            data_start_idx = idx
            break
    if data_start_idx is None:
        confidence = "low"
        return Region(header_idx + 1, header_rows, None, None, [], [], [], confidence, evidence)

    included: list[int] = []
    skipped: list[dict[str, Any]] = []
    exclusion_candidates: list[dict[str, Any]] = []
    blank_streak = 0
    data_end_idx = data_start_idx
    for idx in range(data_start_idx, len(rows)):
        row = rows[idx]
        if is_blank_row(row):
            blank_streak += 1
            if blank_streak >= 2:
                break
            continue
        blank_streak = 0
        candidate = exclusion_candidate(row, header_width)
        if candidate:
            exclusion_candidates.append(
                {
                    **candidate,
                    "row": idx + 1,
                    "row_values": [cell_text(value) for value in row],
                    "content_hash": row_content_hash(row),
                    "decision": "pending",
                }
            )
        included.append(idx + 1)
        data_end_idx = idx

    if score >= 9 and len(included) >= 3:
        confidence = "high"
    elif len(included) >= 1:
        confidence = "medium"
    else:
        confidence = "low"
    return Region(
        header_idx + 1,
        header_rows,
        data_start_idx + 1,
        data_end_idx + 1,
        included,
        skipped,
        exclusion_candidates,
        confidence,
        evidence,
    )


def unique_headers(row: list[Any]) -> list[str]:
    headers = [cell_text(value) or f"column_{idx + 1}" for idx, value in enumerate(trim_trailing_empty(row))]
    counts: dict[str, int] = {}
    result = []
    for header in headers:
        counts[header] = counts.get(header, 0) + 1
        result.append(header if counts[header] == 1 else f"{header}_{counts[header]}")
    return result


def load_csv_rows(path: Path) -> tuple[str, list[list[Any]]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        sample = handle.read(4096)
        handle.seek(0)
        dialect = csv.Sniffer().sniff(sample, delimiters=",\t;")
        return path.stem, [list(row) for row in csv.reader(handle, dialect=dialect)]


def load_xlsx_sheets(path: Path, sheet_name: str | None) -> list[tuple[str, list[list[Any]]]]:
    if load_workbook is None:
        raise SystemExit("openpyxl is required to clean .xlsx files")
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheets = []
    worksheets = [workbook[sheet_name]] if sheet_name else [ws for ws in workbook.worksheets if ws.sheet_state == "visible"]
    for worksheet in worksheets:
        rows = [list(row) for row in worksheet.iter_rows(values_only=True)]
        sheets.append((worksheet.title, rows))
    return sheets


def load_sources(paths: list[Path], sheet_name: str | None) -> list[dict[str, Any]]:
    sources = []
    for path in paths:
        if path.name.startswith(".~"):
            continue
        suffix = path.suffix.lower()
        if suffix == ".xlsx":
            sheet_rows = load_xlsx_sheets(path, sheet_name)
        elif suffix in {".csv", ".tsv"}:
            sheet_rows = [load_csv_rows(path)]
        else:
            raise SystemExit(f"Unsupported file type: {path}")
        for sheet, rows in sheet_rows:
            sources.append({"path": path, "sheet": sheet, "rows": rows})
    return sources


def candidate_id(source_file: str, source_sheet: str, row: int, content_hash: str) -> str:
    payload = f"{source_file}\0{source_sheet}\0{row}\0{content_hash}"
    return f"exclude-{hashlib.sha256(payload.encode('utf-8')).hexdigest()[:16]}"


def bind_candidate_identity(
    candidate: dict[str, Any],
    source_file: str,
    source_sheet: str,
) -> dict[str, Any]:
    result = dict(candidate)
    result["id"] = candidate_id(source_file, source_sheet, int(candidate["row"]), candidate["content_hash"])
    result["source_file"] = source_file
    result["source_sheet"] = source_sheet
    return result


def candidate_counts(candidates: list[dict[str, Any]]) -> dict[str, int]:
    decisions = [
        decision if decision in {"exclude", "keep"} else "pending"
        for candidate in candidates
        for decision in [candidate.get("decision", "pending")]
    ]
    return {
        "total": len(candidates),
        "pending": decisions.count("pending"),
        "exclude": decisions.count("exclude"),
        "keep": decisions.count("keep"),
    }


def build_item(source: dict[str, Any]) -> dict[str, Any]:
    rows = source["rows"]
    region = detect_region(rows)
    if region is None:
        return {
            "source_file": str(source["path"]),
            "source_sheet": source["sheet"],
            "row_count": len(rows),
            "detected": False,
            "confidence": "low",
            "blocking_reasons": ["no_stable_header_detected"],
        }
    header_values = rows[region.header_row - 1]
    headers = unique_headers(header_values)
    source_file = str(source["path"])
    source_sheet = source["sheet"]
    candidates = [
        bind_candidate_identity(candidate, source_file, source_sheet)
        for candidate in region.exclusion_candidates
    ]
    return {
        "source_file": source_file,
        "source_sheet": source_sheet,
        "row_count": len(rows),
        "detected": True,
        "confidence": region.confidence,
        "header_row": region.header_row,
        "header_rows": region.header_rows,
        "data_start_row": region.data_start_row,
        "data_end_row": region.data_end_row,
        "included_row_count": len(region.included_rows),
        "included_rows": region.included_rows,
        "skipped_rows": region.skipped_rows,
        "exclusion_candidates": candidates,
        "exclusion_candidate_counts": candidate_counts(candidates),
        "headers": headers,
        "evidence": region.evidence,
    }


def build_item_from_rule(source: dict[str, Any], rule_item: dict[str, Any]) -> dict[str, Any]:
    rows = source["rows"]
    header_row = int(rule_item["header_row"])
    header_rows = [int(row) for row in rule_item.get("header_rows", [header_row])]
    data_start = int(rule_item["data_start_row"])
    data_end = int(rule_item.get("data_end_row") or len(rows))
    header_values = rows[header_row - 1]
    headers = unique_headers(header_values)
    header_width = len(headers)
    included_rows = []
    skipped_rows = []
    source_file = str(source["path"])
    source_sheet = source["sheet"]
    saved_candidates = {
        int(candidate["row"]): candidate
        for candidate in rule_item.get("exclusion_candidates", [])
        if candidate.get("row") is not None
    }
    candidates = []
    validation_errors = []
    processed_saved_rows = set()
    for row_number in range(data_start, min(data_end, len(rows)) + 1):
        row = rows[row_number - 1]
        saved = saved_candidates.get(row_number)
        if saved:
            processed_saved_rows.add(row_number)
        if is_blank_row(row):
            if saved:
                current_hash = row_content_hash(row)
                validation_errors.append(
                    {
                        "row": row_number,
                        "reason": "candidate_source_row_changed",
                        "expected_hash": saved.get("content_hash"),
                        "actual_hash": current_hash,
                    }
                )
                candidates.append(
                    bind_candidate_identity(
                        {
                            **saved,
                            "row": row_number,
                            "row_values": [cell_text(value) for value in row],
                            "content_hash": current_hash,
                            "decision": "pending",
                        },
                        source_file,
                        source_sheet,
                    )
                )
            continue
        detected = exclusion_candidate(row, header_width)
        if detected or saved:
            current_hash = row_content_hash(row)
            if saved and saved.get("content_hash") != current_hash:
                validation_errors.append(
                    {
                        "row": row_number,
                        "reason": "candidate_source_row_changed",
                        "expected_hash": saved.get("content_hash"),
                        "actual_hash": current_hash,
                    }
                )
            candidate = {
                **(detected or {}),
                **(saved or {}),
                "row": row_number,
                "row_values": [cell_text(value) for value in row],
                "content_hash": current_hash,
            }
            if not saved or saved.get("content_hash") != current_hash:
                candidate["decision"] = "pending"
            candidate = bind_candidate_identity(candidate, source_file, source_sheet)
            candidates.append(candidate)
            if candidate.get("decision") == "exclude":
                skipped_rows.append(
                    {
                        **candidate,
                        "reason": candidate.get("reason", "confirmed_exclusion"),
                        "decision": "exclude",
                    }
                )
                continue
        included_rows.append(row_number)
    for row_number, saved in saved_candidates.items():
        if row_number in processed_saved_rows:
            continue
        validation_errors.append(
            {
                "row": row_number,
                "reason": "candidate_source_row_missing",
                "expected_hash": saved.get("content_hash"),
                "actual_hash": None,
            }
        )
        candidates.append(
            bind_candidate_identity(
                {**saved, "decision": "pending"},
                source_file,
                source_sheet,
            )
        )
    return {
        "source_file": source_file,
        "source_sheet": source_sheet,
        "row_count": len(rows),
        "detected": True,
        "confidence": rule_item.get("confidence", "manual"),
        "header_row": header_row,
        "header_rows": header_rows,
        "data_start_row": data_start,
        "data_end_row": data_end,
        "included_row_count": len(included_rows),
        "included_rows": included_rows,
        "skipped_rows": skipped_rows,
        "exclusion_candidates": candidates,
        "exclusion_candidate_counts": candidate_counts(candidates),
        "validation_errors": validation_errors,
        "headers": headers,
        "evidence": ["loaded from confirmed rules JSON"],
    }


def build_profile_from_rules(sources: list[dict[str, Any]], rules_path: Path) -> dict[str, Any]:
    rules = json.loads(rules_path.read_text(encoding="utf-8"))
    rule_items = rules.get("items", [])
    items = []
    for source in sources:
        exact = None
        fallback = None
        for item in rule_items:
            same_sheet = item.get("source_sheet") == source["sheet"]
            if not same_sheet:
                continue
            if item.get("source_file") == str(source["path"]):
                exact = item
                break
            if Path(str(item.get("source_file", ""))).name == source["path"].name:
                fallback = item
        rule_item = exact or fallback
        if not rule_item:
            items.append(
                {
                    "source_file": str(source["path"]),
                    "source_sheet": source["sheet"],
                    "row_count": len(source["rows"]),
                    "detected": False,
                    "confidence": "low",
                    "blocking_reasons": ["no_matching_rule_item"],
                }
            )
            continue
        items.append(build_item_from_rule(source, rule_item))
    return {
        "generated_at": datetime.now().astimezone().isoformat(),
        "input_paths": sorted({str(source["path"]) for source in sources}),
        "rules_path": str(rules_path),
        "items": items,
    }


def default_output_path(paths: list[Path]) -> Path:
    if len(paths) == 1:
        path = paths[0]
        return path.with_name(f"{path.stem}_清洗后.xlsx")
    parent = paths[0].parent
    return parent / "原始数据_合并清洗后.xlsx"


def build_rules(
    profile: dict[str, Any],
    target_sheet: str,
    add_lineage: bool,
    merge_mode: str,
) -> dict[str, Any]:
    return {
        "version": "2.0",
        "generated_at": datetime.now().astimezone().isoformat(),
        "header": {
            "detect_mode": "score_based",
            "use_primary_header_row": True,
            "skip_bilingual_or_system_header": True,
        },
        "data_region": {
            "start_after_header_rows": True,
            "end_on_two_blank_rows": True,
            "exclusion_mode": "confirmed_candidates_only",
            "require_candidate_decisions": True,
        },
        "output": {
            "target_sheet": target_sheet,
            "same_sheet_merge_mode": merge_mode,
            "write_new_file_only": True,
            "preserve_values": True,
            "lineage_columns": ["源文件名", "源工作表名", "源行号"] if add_lineage else [],
        },
        "items": [
            {
                "source_file": item["source_file"],
                "source_sheet": item["source_sheet"],
                "header_row": item.get("header_row"),
                "header_rows": item.get("header_rows", []),
                "data_start_row": item.get("data_start_row"),
                "data_end_row": item.get("data_end_row"),
                "confidence": item.get("confidence"),
                "exclusion_candidates": item.get("exclusion_candidates", []),
            }
            for item in profile.get("items", [])
        ],
    }


def compare_headers(reference: list[str], candidate: list[str]) -> dict[str, Any]:
    max_len = max(len(reference), len(candidate))
    differences = []
    for idx in range(max_len):
        old = reference[idx] if idx < len(reference) else None
        new = candidate[idx] if idx < len(candidate) else None
        if old != new:
            differences.append({"column": idx + 1, "reference": old, "candidate": new})
    return {
        "consistent": not differences,
        "reference_column_count": len(reference),
        "candidate_column_count": len(candidate),
        "differences": differences,
    }


def build_sheet_groups(profile: dict[str, Any]) -> list[dict[str, Any]]:
    grouped: dict[str, list[dict[str, Any]]] = {}
    for item in profile.get("items", []):
        if not item.get("detected"):
            continue
        grouped.setdefault(item["source_sheet"], []).append(item)

    groups = []
    for sheet_name, items in sorted(grouped.items()):
        reference = items[0].get("headers", [])
        sources = []
        all_consistent = True
        for item in items:
            comparison = compare_headers(reference, item.get("headers", []))
            all_consistent = all_consistent and comparison["consistent"]
            sources.append(
                {
                    "source_file": item["source_file"],
                    "source_sheet": item["source_sheet"],
                    "header_row": item.get("header_row"),
                    "header_column_count": len(item.get("headers", [])),
                    "header_consistent": comparison["consistent"],
                    "header_differences": comparison["differences"],
                    "actual_data_rows": item.get("included_row_count", 0),
                    "skipped_row_count": len(item.get("skipped_rows", [])),
                    "exclusion_candidate_counts": item.get("exclusion_candidate_counts", {}),
                }
            )
        groups.append(
            {
                "sheet_name": sheet_name,
                "source_count": len(items),
                "same_name_across_files": len({item["source_file"] for item in items}) > 1,
                "headers_consistent": all_consistent,
                "merge_candidate": len(items) > 1 and all_consistent,
                "requires_merge_decision": len(items) > 1 and all_consistent,
                "sources": sources,
                "merged_data_rows": sum(item.get("included_row_count", 0) for item in items),
                "merged_skipped_rows": sum(len(item.get("skipped_rows", [])) for item in items),
                "merged_exclusion_candidates": sum(
                    item.get("exclusion_candidate_counts", {}).get("total", 0) for item in items
                ),
            }
        )
    return groups


def structure_issues(profile: dict[str, Any]) -> list[dict[str, Any]]:
    issues = []
    for group in profile.get("sheet_groups", []):
        if group["headers_consistent"]:
            continue
        issues.append(
            {
                "type": "same_sheet_incompatible_headers",
                "sheet_name": group["sheet_name"],
                "evidence": f"同名 Tab「{group['sheet_name']}」在不同文件中的表头不一致，不能直接合并。",
                "sources": group["sources"],
            }
        )
    return issues


def safe_sheet_name(name: str, existing: set[str]) -> str:
    cleaned = re.sub(r"[:\\/?*\[\]]", "_", name).strip() or "Sheet"
    cleaned = cleaned[:31]
    candidate = cleaned
    index = 2
    while candidate in existing:
        suffix = f"_{index}"
        candidate = f"{cleaned[:31-len(suffix)]}{suffix}"
        index += 1
    existing.add(candidate)
    return candidate


def output_groups(
    profile: dict[str, Any],
    target_sheet: str,
    merge_same_sheets: bool,
    keep_separate_sheets: bool,
) -> list[dict[str, Any]]:
    items = [item for item in profile["items"] if item.get("detected") and item.get("included_rows")]
    if len(items) == 1:
        return [{"output_sheet": target_sheet, "items": items}]

    grouped: dict[str, list[dict[str, Any]]] = {}
    for item in items:
        grouped.setdefault(item["source_sheet"], []).append(item)

    result = []
    for sheet_name, sheet_items in sorted(grouped.items()):
        if len(sheet_items) == 1 or merge_same_sheets:
            result.append({"output_sheet": sheet_name, "items": sheet_items})
            continue
        if keep_separate_sheets:
            for item in sheet_items:
                result.append(
                    {
                        "output_sheet": f"{sheet_name}_{Path(item['source_file']).stem}",
                        "items": [item],
                    }
                )
    return result


def write_output(
    profile: dict[str, Any],
    sources: list[dict[str, Any]],
    output_path: Path,
    target_sheet: str,
    add_lineage: bool,
    merge_same_sheets: bool,
    keep_separate_sheets: bool,
) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    if Workbook is None:
        raise SystemExit("openpyxl is required to write .xlsx files")
    source_key = {(str(source["path"]), source["sheet"]): source for source in sources}
    workbook = Workbook()
    workbook.remove(workbook.active)
    existing_names: set[str] = set()
    written_groups = []
    for group in output_groups(profile, target_sheet, merge_same_sheets, keep_separate_sheets):
        output_sheet = safe_sheet_name(group["output_sheet"], existing_names)
        worksheet = workbook.create_sheet(output_sheet)
        first_item = group["items"][0]
        canonical_headers = list(first_item["headers"])
        output_headers = list(canonical_headers)
        if add_lineage:
            output_headers.extend(["源文件名", "源工作表名", "源行号"])
        worksheet.append(output_headers)

        source_stats = []
        for item in group["items"]:
            source = source_key[(item["source_file"], item["source_sheet"])]
            rows = source["rows"]
            for row_number in item["included_rows"]:
                row = rows[row_number - 1]
                values = list(row[: len(canonical_headers)])
                if len(values) < len(canonical_headers):
                    values.extend([None] * (len(canonical_headers) - len(values)))
                if add_lineage:
                    values.extend([Path(item["source_file"]).name, item["source_sheet"], row_number])
                worksheet.append(values)
            source_stats.append(
                {
                    "source_file": item["source_file"],
                    "source_sheet": item["source_sheet"],
                    "actual_data_rows": item["included_row_count"],
                    "excluded_rows": len(item.get("skipped_rows", [])),
                }
            )
        written_groups.append(
            {
                "output_sheet": output_sheet,
                "source_count": len(group["items"]),
                "source_stats": source_stats,
                "total_data_rows": sum(item["included_row_count"] for item in group["items"]),
                "total_excluded_rows": sum(len(item.get("skipped_rows", [])) for item in group["items"]),
            }
        )

    audit_sheet_name = safe_sheet_name(AUDIT_SHEET_NAME, existing_names)
    audit_sheet = workbook.create_sheet(audit_sheet_name)
    audit_sheet.append(
        [
            "来源文件",
            "来源工作表",
            "源行号",
            "排除类型",
            "命中内容",
            "排除依据",
            "用户决策",
            "原始行内容",
        ]
    )
    audit_count = 0
    for item in profile["items"]:
        for skipped in item.get("skipped_rows", []):
            matched = json.dumps(skipped.get("matched_cells", []), ensure_ascii=False)
            row_values = json.dumps(skipped.get("row_values", []), ensure_ascii=False)
            audit_sheet.append(
                [
                    item["source_file"],
                    item["source_sheet"],
                    skipped["row"],
                    skipped.get("category", skipped.get("reason", "")),
                    matched,
                    skipped.get("evidence", ""),
                    skipped.get("decision", "exclude"),
                    row_values,
                ]
            )
            audit_count += 1
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    return written_groups, {"sheet_name": audit_sheet_name, "excluded_row_count": audit_count}


def main() -> None:
    args = parse_args()
    input_paths = [Path(item).expanduser().resolve() for item in args.inputs]
    output_path = Path(args.output).expanduser().resolve() if args.output else default_output_path(input_paths)
    if any(output_path == path for path in input_paths):
        raise SystemExit("Output path must be different from every input path; original files are read-only.")

    sources = load_sources(input_paths, args.sheet)
    if args.rules:
        profile = build_profile_from_rules(sources, Path(args.rules).expanduser().resolve())
    else:
        profile = {
            "generated_at": datetime.now().astimezone().isoformat(),
            "input_paths": [str(path) for path in input_paths],
            "items": [build_item(source) for source in sources],
        }
    profile["sheet_groups"] = build_sheet_groups(profile)
    profile["blocking_issues"] = structure_issues(profile)
    add_lineage = not args.no_lineage
    requested_merge_mode = (
        "merge"
        if args.merge_same_sheets
        else "keep_separate"
        if args.keep_separate_sheets
        else "pending_confirmation"
    )
    rules = (
        json.loads(Path(args.rules).expanduser().resolve().read_text(encoding="utf-8"))
        if args.rules
        else build_rules(profile, args.target_sheet, add_lineage, requested_merge_mode)
    )
    saved_merge_mode = rules.get("output", {}).get("same_sheet_merge_mode")
    merge_mode = requested_merge_mode if requested_merge_mode != "pending_confirmation" else saved_merge_mode
    merge_mode = merge_mode or "pending_confirmation"
    rules.setdefault("output", {})["same_sheet_merge_mode"] = merge_mode
    merge_candidates = [group for group in profile["sheet_groups"] if group["requires_merge_decision"]]
    merge_decision_required = bool(merge_candidates) and merge_mode == "pending_confirmation"
    blocking = [item for item in profile["items"] if not item.get("detected") or item.get("confidence") == "low"]
    all_candidates = [
        candidate
        for item in profile["items"]
        for candidate in item.get("exclusion_candidates", [])
    ]
    pending_candidates = [
        candidate
        for candidate in all_candidates
        if candidate.get("decision") not in {"exclude", "keep"}
    ]
    candidate_validation_errors = [
        {
            "source_file": item["source_file"],
            "source_sheet": item["source_sheet"],
            **error,
        }
        for item in profile["items"]
        for error in item.get("validation_errors", [])
    ]
    exclusion_counts = candidate_counts(all_candidates)
    summary = {
        "generated_at": datetime.now().astimezone().isoformat(),
        "mode": "execute" if args.execute else "dry_run",
        "output_path": str(output_path),
        "target_sheet": args.target_sheet,
        "source_count": len(sources),
        "total_included_rows": sum(item.get("included_row_count", 0) for item in profile["items"]),
        "blocking_count": (
            len(blocking)
            + len(profile["blocking_issues"])
            + int(bool(pending_candidates))
            + int(bool(candidate_validation_errors))
        ),
        "blocking_issues": profile["blocking_issues"],
        "exclusion_candidate_counts": exclusion_counts,
        "exclusion_confirmation_required": bool(pending_candidates),
        "candidate_validation_errors": candidate_validation_errors,
        "merge_mode": merge_mode,
        "merge_decision_required": merge_decision_required,
        "sheet_groups": profile["sheet_groups"],
        "items": [
            {
                "source_file": item["source_file"],
                "source_sheet": item["source_sheet"],
                "confidence": item.get("confidence"),
                "header_rows": item.get("header_rows", []),
                "data_start_row": item.get("data_start_row"),
                "data_end_row": item.get("data_end_row"),
                "included_row_count": item.get("included_row_count", 0),
                "skipped_rows": item.get("skipped_rows", []),
                "exclusion_candidates": item.get("exclusion_candidates", []),
                "exclusion_candidate_counts": item.get("exclusion_candidate_counts", {}),
                "validation_errors": item.get("validation_errors", []),
            }
            for item in profile["items"]
        ],
    }

    if args.profile_output:
        Path(args.profile_output).write_text(json.dumps(profile, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    if args.rules_output:
        Path(args.rules_output).write_text(json.dumps(rules, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    if args.run_output:
        Path(args.run_output).write_text(json.dumps(summary, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    if args.execute:
        if blocking or profile["blocking_issues"]:
            raise SystemExit("Blocking structure issues remain; review dry-run output before executing.")
        if candidate_validation_errors:
            raise SystemExit("排除候选对应的源行已变化；请重新 dry-run 并让用户确认新的候选行。")
        if pending_candidates:
            rows = ", ".join(
                f"{Path(candidate['source_file']).name}/{candidate['source_sheet']}:{candidate['row']}"
                for candidate in pending_candidates[:10]
            )
            raise SystemExit(f"仍有待用户确认的排除候选行，不能执行清洗：{rows}")
        if merge_decision_required:
            names = ", ".join(group["sheet_name"] for group in merge_candidates)
            raise SystemExit(
                f"以下同名 Tab 需要用户确认是否合并：{names}。"
                "用户确认合并后使用 --merge-same-sheets；确认分别保留后使用 --keep-separate-sheets。"
            )
        if args.rules:
            Path(args.rules).expanduser().resolve().write_text(
                json.dumps(rules, ensure_ascii=False, indent=2) + "\n",
                encoding="utf-8",
            )
        written_groups, exclusion_audit = write_output(
            profile,
            sources,
            output_path,
            args.target_sheet,
            add_lineage,
            merge_mode == "merge",
            merge_mode == "keep_separate",
        )
        summary["written"] = True
        summary["written_sheet_groups"] = written_groups
        summary["exclusion_audit"] = exclusion_audit
        if args.handoff_output:
            handoff_path = Path(args.handoff_output).expanduser().resolve()
            if args.rules:
                confirmed_rules_path = Path(args.rules).expanduser().resolve()
            elif args.rules_output:
                confirmed_rules_path = Path(args.rules_output).expanduser().resolve()
            else:
                confirmed_rules_path = handoff_path.parent / "rules.json"
                confirmed_rules_path.parent.mkdir(parents=True, exist_ok=True)
                confirmed_rules_path.write_text(
                    json.dumps(rules, ensure_ascii=False, indent=2) + "\n",
                    encoding="utf-8",
                )
            handoff = {
                "schema_version": "1.1",
                "generated_at": datetime.now().astimezone().isoformat(),
                "cleaning_status": "completed",
                "analysis_gate": {
                    "status": "awaiting_user_confirmation",
                    "confirmed_at": None,
                },
                "analysis_goal_gate": {
                    "status": "awaiting_confirmation",
                    "confirmed_at": None,
                    "goal": None,
                    "decision_object": None,
                    "focus": None,
                    "output_depth": None,
                    "visualization_mode": "自动判定",
                    "report_format": "Markdown + HTML",
                    "business_context": None,
                    "analysis_sheets": [],
                },
                "cleaned_file_path": str(output_path),
                "sheet_name": args.target_sheet,
                "cleaning_run_id": args.cleaning_run_id,
                "rules": {
                    "status": "saved",
                    "path": str(confirmed_rules_path),
                    "description": "本轮清洗实际采用的表头、数据范围和排除规则",
                },
                "rules_path": str(confirmed_rules_path),
                "source_files": [str(path) for path in input_paths],
                "lineage_columns": rules.get("output", {}).get("lineage_columns", []),
                "output_sheets": written_groups,
                "exclusion_audit": exclusion_audit,
                "excluded_rows": [
                    {
                        "source_file": item["source_file"],
                        "source_sheet": item["source_sheet"],
                        **skipped,
                    }
                    for item in profile["items"]
                    for skipped in item.get("skipped_rows", [])
                ],
                "warnings": [
                    {
                        "source_file": item["source_file"],
                        "source_sheet": item["source_sheet"],
                        "skipped_rows": item.get("skipped_rows", []),
                    }
                    for item in profile["items"]
                    if item.get("skipped_rows")
                ],
                "profile_summary": {
                    "source_count": len(sources),
                    "output_row_count": summary["total_included_rows"],
                    "sheet_groups": profile["sheet_groups"],
                    "written_sheet_groups": written_groups,
                    "items": summary["items"],
                },
            }
            if len(written_groups) == 1:
                handoff["sheet_name"] = written_groups[0]["output_sheet"]
            handoff_path.write_text(
                json.dumps(handoff, ensure_ascii=False, indent=2) + "\n",
                encoding="utf-8",
            )
            summary["handoff_path"] = str(handoff_path)
            summary["next_step"] = "stop_and_wait_for_user_confirmation"
        if args.run_output:
            Path(args.run_output).write_text(json.dumps(summary, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
