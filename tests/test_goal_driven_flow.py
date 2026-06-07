from __future__ import annotations

import csv
import json
import subprocess
import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
SCRIPTS = ROOT / "scripts"


def run_script(name: str, *args: str) -> dict:
    completed = subprocess.run(
        ["python3", str(SCRIPTS / name), *args],
        check=False,
        capture_output=True,
        text=True,
    )
    if completed.returncode != 0:
        raise AssertionError(
            f"{name} failed with exit code {completed.returncode}\n"
            f"stdout:\n{completed.stdout}\nstderr:\n{completed.stderr}"
        )
    return json.loads(completed.stdout)


class GoalDrivenFlowTest(unittest.TestCase):
    def test_legacy_mapping_without_source_file_blocks_when_ambiguous(self) -> None:
        with tempfile.TemporaryDirectory() as temporary_directory:
            workdir = Path(temporary_directory)
            first_dir = workdir / "first"
            second_dir = workdir / "second"
            first_dir.mkdir()
            second_dir.mkdir()
            inputs = [first_dir / "sales.csv", second_dir / "sales.csv"]
            for path in inputs:
                with path.open("w", encoding="utf-8-sig", newline="") as handle:
                    writer = csv.writer(handle)
                    writer.writerows(
                        [
                            ["Revenue", "Customer"],
                            [100, "A"],
                            [120, "B"],
                            [80, "C"],
                        ]
                    )

            contract = {
                "status": "confirmed",
                "goal_id": "goal-ambiguous",
                "goal": "分析收入",
                "contract_fingerprint": "sha256:test",
                "required_data": {
                    "scope": [
                        {"file": str(inputs[0].resolve()), "sheet": "sales"},
                        {"file": str(inputs[1].resolve()), "sheet": "sales"},
                    ],
                    "required_fields": ["Revenue"],
                    "supporting_fields": [],
                    "join_keys": [],
                    "time_fields": [],
                },
            }
            mapping = {
                "status": "confirmed",
                "contract_fingerprint": "sha256:test",
                "mappings": [
                    {
                        "sheet": "sales",
                        "source_field": "Revenue",
                        "goal_field": "Revenue",
                    }
                ],
            }
            contract_path = workdir / "goal-contract.json"
            mapping_path = workdir / "field-mapping.json"
            contract_path.write_text(json.dumps(contract, ensure_ascii=False), encoding="utf-8")
            mapping_path.write_text(json.dumps(mapping, ensure_ascii=False), encoding="utf-8")

            completed = subprocess.run(
                [
                    "python3",
                    str(SCRIPTS / "clean_tabular_data.py"),
                    *(str(path) for path in inputs),
                    "--goal-contract",
                    str(contract_path),
                    "--field-mapping",
                    str(mapping_path),
                ],
                check=False,
                capture_output=True,
                text=True,
            )
            self.assertNotEqual(completed.returncode, 0)
            self.assertIn("ambiguous_mapping_source", completed.stderr)

    def test_misaligned_sources_generate_independent_targeted_file(self) -> None:
        with tempfile.TemporaryDirectory() as temporary_directory:
            workdir = Path(temporary_directory)
            first = workdir / "january.csv"
            second = workdir / "february.csv"
            with first.open("w", encoding="utf-8-sig", newline="") as handle:
                writer = csv.writer(handle)
                writer.writerows(
                    [
                        ["Date", "Revenue", "Customer", "Unused"],
                        ["2026-01-01", 100, "A", "x"],
                        ["2026-01-02", 120, "B", "y"],
                        ["2026-01-03", 80, "A", "z"],
                    ]
                )
            with second.open("w", encoding="utf-8-sig", newline="") as handle:
                writer = csv.writer(handle)
                writer.writerows(
                    [
                        ["日期", "收入", "客户", "无关列"],
                        ["2026-02-01", 90, "A", "x"],
                        ["2026-02-02", "", "C", "y"],
                        ["2026-02-03", 130, "C", "z"],
                    ]
                )

            data_session = workdir / ".data-session"
            cleaning_session = workdir / ".cleaning-session"
            discovery = data_session / "discovery.json"
            run_script(
                "profile_dataset.py",
                str(first),
                str(second),
                "--all-sheets",
                "--output",
                str(discovery),
            )
            discovery_payload = json.loads(discovery.read_text(encoding="utf-8"))
            self.assertEqual(discovery_payload["mode"], "discovery")
            self.assertTrue(discovery_payload["goal_candidates"])

            run_script(
                "data_session_store.py",
                "init",
                "--session-dir",
                str(data_session),
                "--discovery",
                str(discovery),
            )
            scope = json.dumps(
                [
                    {"file": str(first.resolve()), "sheet": "january"},
                    {"file": str(second.resolve()), "sheet": "february"},
                ],
                ensure_ascii=False,
            )
            contract = run_script(
                "data_session_store.py",
                "confirm-goal",
                "--session-dir",
                str(data_session),
                "--goal",
                "分析月度收入及客户贡献",
                "--goal-id",
                "goal-revenue",
                "--goal-type",
                "business_analysis",
                "--questions",
                '["收入如何变化","哪些客户贡献最高"]',
                "--scope",
                scope,
                "--required-fields",
                "Revenue",
                "--supporting-fields",
                "Customer",
                "--time-fields",
                "Date",
            )
            mappings = [
                {
                    "sheet": "february",
                    "source_field": "收入",
                    "goal_field": "Revenue",
                    "target_sheet": "目标数据",
                },
                {
                    "sheet": "february",
                    "source_field": "客户",
                    "goal_field": "Customer",
                    "target_sheet": "目标数据",
                },
                {
                    "sheet": "february",
                    "source_field": "日期",
                    "goal_field": "Date",
                    "target_sheet": "目标数据",
                },
                {
                    "source_file": str(first.resolve()),
                    "source_sheet": "january",
                    "source_field": "Revenue",
                    "target_field": "Revenue",
                    "target_sheet": "目标数据",
                },
            ]
            mapping_path = workdir / "mappings.json"
            mapping_path.write_text(json.dumps(mappings, ensure_ascii=False), encoding="utf-8")
            run_script(
                "data_session_store.py",
                "confirm-mapping",
                "--session-dir",
                str(data_session),
                "--mappings",
                str(mapping_path),
            )

            profile = workdir / "profile.json"
            quality = workdir / "quality.json"
            run_script("profile_dataset.py", str(first), "--output", str(profile))
            run_script(
                "detect_anomalies.py",
                str(first),
                "--profile",
                str(profile),
                "--goal-contract",
                str(data_session / "goal-contract.json"),
                "--field-mapping",
                str(data_session / "field-mapping.json"),
                "--output",
                str(quality),
            )
            run_script(
                "data_session_store.py",
                "save-quality",
                "--session-dir",
                str(data_session),
                "--input",
                str(quality),
            )
            second_profile = workdir / "second-profile.json"
            second_quality = workdir / "second-quality.json"
            run_script("profile_dataset.py", str(second), "--output", str(second_profile))
            run_script(
                "detect_anomalies.py",
                str(second),
                "--profile",
                str(second_profile),
                "--goal-contract",
                str(data_session / "goal-contract.json"),
                "--field-mapping",
                str(data_session / "field-mapping.json"),
                "--output",
                str(second_quality),
            )
            run_script(
                "data_session_store.py",
                "save-quality",
                "--session-dir",
                str(data_session),
                "--input",
                str(second_quality),
            )
            quality_payload = json.loads((data_session / "quality-impact.json").read_text(encoding="utf-8"))
            mapped_missing = [
                item
                for item in quality_payload["anomalies"]
                if item.get("details", {}).get("source_field_name") == "收入"
            ]
            self.assertTrue(mapped_missing)
            self.assertEqual(mapped_missing[0]["affected_fields"], ["Revenue"])
            self.assertEqual(mapped_missing[0]["impact_level"], "blocking")
            anomaly_ids = [item["id"] for item in quality_payload["anomalies"]]
            if anomaly_ids:
                run_script(
                    "data_session_store.py",
                    "decide-quality",
                    "--session-dir",
                    str(data_session),
                    "--target-id",
                    *anomaly_ids,
                    "--choice",
                    "accept",
                )

            output = workdir / "revenue_targeted.xlsx"
            rules = cleaning_session / "rules.json"
            run_summary = cleaning_session / "dry-run.json"
            run_script(
                "clean_tabular_data.py",
                str(first),
                str(second),
                "--output",
                str(output),
                "--goal-contract",
                str(data_session / "goal-contract.json"),
                "--field-mapping",
                str(data_session / "field-mapping.json"),
                "--profile-output",
                str(cleaning_session / "profile.json"),
                "--rules-output",
                str(rules),
                "--run-output",
                str(run_summary),
            )
            run_script(
                "clean_tabular_data.py",
                str(first),
                str(second),
                "--output",
                str(output),
                "--goal-contract",
                str(data_session / "goal-contract.json"),
                "--field-mapping",
                str(data_session / "field-mapping.json"),
                "--quality-impact",
                str(data_session / "quality-impact.json"),
                "--rules",
                str(rules),
                "--run-output",
                str(cleaning_session / "run-summary.json"),
                "--handoff-output",
                str(cleaning_session / "handoff.json"),
                "--cleaning-run-id",
                "run-001",
                "--execute",
            )

            workbook = load_workbook(output, read_only=True, data_only=True)
            self.assertIn("数据说明", workbook.sheetnames)
            self.assertIn("清洗审计", workbook.sheetnames)
            data_sheets = [name for name in workbook.sheetnames if name not in {"数据说明", "清洗审计"}]
            self.assertEqual(data_sheets, ["目标数据"])
            for sheet_name in data_sheets:
                headers = [cell.value for cell in next(workbook[sheet_name].iter_rows())]
                self.assertEqual(
                    headers,
                    ["Revenue", "Customer", "Date", "源文件名", "源工作表名", "源行号"],
                )
                self.assertNotIn("Unused", headers)
                self.assertNotIn("无关列", headers)

            audit_rows = list(workbook["清洗审计"].iter_rows(values_only=True))
            self.assertEqual(
                list(audit_rows[0]),
                [
                    "审计类型",
                    "来源文件",
                    "来源工作表",
                    "来源表头行",
                    "源数据行号",
                    "来源字段 / 命中内容",
                    "目标字段 / 处理依据",
                    "目标工作表",
                    "用户决策",
                    "原始内容 / 记录 ID",
                ],
            )
            mapping_rows = [row for row in audit_rows[1:] if row[0] == "字段映射"]
            self.assertEqual(len(mapping_rows), len(mappings))
            for row in mapping_rows:
                self.assertTrue(row[1])
                self.assertTrue(row[2])
                self.assertEqual(row[3], 1)
                self.assertEqual(row[4], "不适用")
                self.assertTrue(row[5])
                self.assertTrue(row[6])
                self.assertEqual(row[7], "目标数据")

            handoff = json.loads((cleaning_session / "handoff.json").read_text(encoding="utf-8"))
            self.assertEqual(handoff["goal_id"], contract["goal_id"])
            self.assertEqual(handoff["targeted_cleaning_file_path"], str(output.resolve()))
            self.assertEqual(handoff["analysis_goal_gate"]["status"], "confirmed")
            self.assertEqual(handoff["field_mapping"]["schema_version"], "2.0")
            for mapping in handoff["field_mapping"]["mappings"]:
                self.assertTrue(mapping["source_file"])
                self.assertTrue(mapping["source_sheet"])
                self.assertEqual(mapping["source_header_row"], 1)
                self.assertTrue(mapping["source_field"])
                self.assertTrue(mapping["target_field"])
                self.assertEqual(mapping["target_sheet"], "目标数据")
            self.assertEqual(handoff["original_field_mapping"]["mappings"], mappings)
            analysis_session = workdir / ".analysis-session"
            analysis_state = run_script(
                "session_store.py",
                "init",
                "--session-dir",
                str(analysis_session),
                "--dataset",
                str(output),
                "--goal",
                "分析月度收入及客户贡献",
                "--goal-confirmed",
                "--goal-contract",
                str(data_session / "goal-contract.json"),
                "--analysis-sheets",
                "目标数据",
            )
            self.assertEqual(analysis_state["goal_id"], "goal-revenue")
            self.assertEqual(analysis_state["dataset_path"], str(output.resolve()))


if __name__ == "__main__":
    unittest.main()
